import json
import os
import sys
import threading
import time
from datetime import datetime
from typing import List, Dict, Any

import pika
from openpyxl import Workbook, load_workbook

REPORT_FILE = "events_report.xlsx"
QUEUE_NAME = "excel_report_queue"
RABBITMQ_HOST = "localhost"
RABBITMQ_PORT = 5672

BUFFER_SIZE = 10
BUFFER_FLUSH_INTERVAL = 30


class ExcelReportService:
    def __init__(
        self,
        report_file: str = REPORT_FILE,
        queue_name: str = QUEUE_NAME,
        rabbitmq_host: str = RABBITMQ_HOST,
        rabbitmq_port: int = RABBITMQ_PORT,
        buffer_size: int = BUFFER_SIZE,
        buffer_flush_interval: int = BUFFER_FLUSH_INTERVAL,
    ):
        self.report_file = report_file
        self.queue_name = queue_name
        self.rabbitmq_host = rabbitmq_host
        self.rabbitmq_port = rabbitmq_port
        self.buffer_size = buffer_size
        self.buffer_flush_interval = buffer_flush_interval

        self.buffer: List[Dict[str, Any]] = []
        self.buffer_lock = threading.Lock()
        self.flush_timer: threading.Timer | None = None
        self.connection: pika.BlockingConnection | None = None
        self.channel: pika.channel.Channel | None = None

    def _ensure_report_file(self) -> None:
        if os.path.exists(self.report_file):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Events"
        ws.append([
            "ID",
            "Username",
            "Email",
            "Event Type",
            "Timestamp",
            "Details",
        ])
        wb.save(self.report_file)
        print(f"[INFO] Created new report file: {self.report_file}")

    def _write_buffer_to_excel(self) -> None:
        with self.buffer_lock:
            if not self.buffer:
                return
            records = self.buffer[:]
            self.buffer.clear()

        try:
            wb = load_workbook(self.report_file)
            ws = wb.active
            for record in records:
                ws.append([
                    record.get("id", ""),
                    record.get("username", ""),
                    record.get("email", ""),
                    record.get("event_type", ""),
                    record.get("timestamp", datetime.now().isoformat()),
                    record.get("details", ""),
                ])
            wb.save(self.report_file)
            print(f"[INFO] Flushed {len(records)} records to {self.report_file}")
        except Exception as exc:
            print(f"[ERROR] Failed to write to Excel: {exc}")
            with self.buffer_lock:
                self.buffer = records + self.buffer

    def _schedule_flush(self) -> None:
        if self.flush_timer:
            self.flush_timer.cancel()
        self.flush_timer = threading.Timer(
            self.buffer_flush_interval, self._timed_flush
        )
        self.flush_timer.daemon = True
        self.flush_timer.start()

    def _timed_flush(self) -> None:
        self._write_buffer_to_excel()
        self._schedule_flush()

    def _on_message(
        self,
        ch: pika.channel.Channel,
        method: pika.frame.Method,
        properties: pika.BasicProperties,
        body: bytes,
    ) -> None:
        try:
            data = json.loads(body.decode("utf-8"))
            print(f"[RECV] {data}")
        except json.JSONDecodeError as exc:
            print(f"[ERROR] Invalid JSON: {exc}")
            ch.basic_ack(delivery_tag=method.delivery_tag)
            return

        record = {
            "id": data.get("id", ""),
            "username": data.get("username", ""),
            "email": data.get("email", ""),
            "event_type": data.get("event_type", ""),
            "timestamp": data.get("timestamp", datetime.now().isoformat()),
            "details": data.get("details", ""),
        }

        with self.buffer_lock:
            self.buffer.append(record)
            should_flush = len(self.buffer) >= self.buffer_size

        ch.basic_ack(delivery_tag=method.delivery_tag)

        if should_flush:
            self._write_buffer_to_excel()
            self._schedule_flush()

    def start(self) -> None:
        self._ensure_report_file()
        self._schedule_flush()

        try:
            self.connection = pika.BlockingConnection(
                pika.ConnectionParameters(
                    host=self.rabbitmq_host,
                    port=self.rabbitmq_port,
                )
            )
            self.channel = self.connection.channel()
            self.channel.queue_declare(queue=self.queue_name, durable=True)
            self.channel.basic_qos(prefetch_count=self.buffer_size)
            self.channel.basic_consume(
                queue=self.queue_name, on_message_callback=self._on_message
            )
            print(
                f"[INFO] Service started. Listening on queue '{self.queue_name}'."
            )
            print(f"[INFO] Buffer size: {self.buffer_size}, flush interval: {self.buffer_flush_interval}s")
            self.channel.start_consuming()
        except pika.exceptions.AMQPConnectionError as exc:
            print(f"[ERROR] RabbitMQ connection failed: {exc}")
            sys.exit(1)
        except KeyboardInterrupt:
            print("\n[INFO] Interrupted by user.")
        finally:
            self.stop()

    def stop(self) -> None:
        if self.flush_timer:
            self.flush_timer.cancel()
        self._write_buffer_to_excel()
        if self.channel and self.channel.is_open:
            self.channel.stop_consuming()
        if self.connection and self.connection.is_open:
            self.connection.close()
        print("[INFO] Service stopped.")


def main() -> None:
    service = ExcelReportService()
    service.start()


if __name__ == "__main__":
    main()
